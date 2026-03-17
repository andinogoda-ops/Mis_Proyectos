import os
import io
import sqlite3
from datetime import datetime, date, timedelta

from flask import (
    Flask,
    render_template_string,
    request,
    redirect,
    url_for,
    flash,
    Response,
    send_file,
)

# ============================================================
# App
# ============================================================
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "cambia-esta-clave-segura")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "pos.db")

# Regalías
REGALIA_CODE_DEFAULT = "SPOT2025"


# ============================================================
# DB helpers
# ============================================================
def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def table_exists(conn: sqlite3.Connection, table: str) -> bool:
    r = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name=?;",
        (table,),
    ).fetchone()
    return r is not None


def column_exists(conn: sqlite3.Connection, table: str, column: str) -> bool:
    if not table_exists(conn, table):
        return False
    cols = conn.execute(f"PRAGMA table_info({table});").fetchall()
    return any(r["name"] == column for r in cols)


def ensure_column(conn: sqlite3.Connection, table: str, column: str, ddl_fragment: str):
    """
    ddl_fragment ejemplo: "payment_method TEXT"
    """
    if not column_exists(conn, table, column):
        conn.execute(f"ALTER TABLE {table} ADD COLUMN {ddl_fragment};")


def init_db():
    conn = get_conn()
    try:
        # products: inventario + costo + venta + imagen BLOB
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS products (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                category TEXT NOT NULL,
                sale_price INTEGER NOT NULL DEFAULT 0,
                cost_price INTEGER NOT NULL DEFAULT 0,
                stock INTEGER NOT NULL DEFAULT 0,
                image_blob BLOB,
                image_mime TEXT,
                created_at TEXT NOT NULL
            );
            """
        )

        # settings: para guardar regalia_code
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS settings (
                k TEXT PRIMARY KEY,
                v TEXT NOT NULL
            );
            """
        )

        # sales: historial (con costo/ganancia + método pago + estado)
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS sales (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                created_at TEXT NOT NULL,
                account_name TEXT NOT NULL,
                total INTEGER NOT NULL DEFAULT 0,
                cost_total INTEGER NOT NULL DEFAULT 0,
                profit INTEGER NOT NULL DEFAULT 0,
                is_gift INTEGER NOT NULL DEFAULT 0,
                cash_received INTEGER,
                change_given INTEGER,
                payment_method TEXT,
                status TEXT NOT NULL DEFAULT 'paid',
                closed_at TEXT
            );
            """
        )

        # sale_items: detalle de cada venta + despacho
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS sale_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                sale_id INTEGER NOT NULL,
                product_id INTEGER,
                name TEXT NOT NULL,
                category TEXT NOT NULL,
                sale_price INTEGER NOT NULL,
                cost_price INTEGER NOT NULL,
                created_at TEXT NOT NULL,
                delivered INTEGER NOT NULL DEFAULT 0,
                delivered_at TEXT,
                FOREIGN KEY(sale_id) REFERENCES sales(id)
            );
            """
        )

        # ---- Migraciones suaves (por si tu DB ya existía) ----
        ensure_column(conn, "sales", "payment_method", "payment_method TEXT")
        ensure_column(conn, "sales", "cash_received", "cash_received INTEGER")
        ensure_column(conn, "sales", "change_given", "change_given INTEGER")
        # Estas 2 a veces ya están, pero igual protegemos
        if not column_exists(conn, "sales", "cost_total"):
            ensure_column(conn, "sales", "cost_total", "cost_total INTEGER NOT NULL DEFAULT 0")
        if not column_exists(conn, "sales", "profit"):
            ensure_column(conn, "sales", "profit", "profit INTEGER NOT NULL DEFAULT 0")

        # Estado de venta (para cocina)
        ensure_column(conn, "sales", "status", "status TEXT NOT NULL DEFAULT 'paid'")
        ensure_column(conn, "sales", "closed_at", "closed_at TEXT")

        # Despacho por ítem
        ensure_column(conn, "sale_items", "delivered", "delivered INTEGER NOT NULL DEFAULT 0")
        ensure_column(conn, "sale_items", "delivered_at", "delivered_at TEXT")

        # settings default
        if conn.execute("SELECT v FROM settings WHERE k='regalia_code';").fetchone() is None:
            conn.execute("INSERT INTO settings(k,v) VALUES('regalia_code', ?);", (REGALIA_CODE_DEFAULT,))

        # seed si está vacío
        count = conn.execute("SELECT COUNT(*) AS c FROM products;").fetchone()["c"]
        if count == 0:
            seed_products(conn)

        conn.commit()
    finally:
        conn.close()


def seed_products(conn: sqlite3.Connection):
    now = datetime.now().isoformat()

    comidas = [
        ("Hamburguesa Especial", 3200, 2200, 50),
        ("Hamburguesa Especial con Papas", 3900, 2700, 50),
        ("Empanadas", 1600, 900, 50),
        ("Empanadas Arregladas", 1900, 1100, 50),
        ("Taco", 1500, 800, 50),
        ("Taco con Papas", 2300, 1400, 50),
        ("Perro Caliente", 1500, 900, 50),
        ("Perro Caliente con Papas", 2300, 1500, 50),
        ("Salchipapas", 1800, 1100, 50),
        ("Papas Fritas", 1300, 700, 50),
        ("Caldosas", 1500, 900, 50),
    ]
    bebidas = [
        ("Gaseosa 600ml", 1000, 650, 50),
        ("Gaseosa 355ml", 600, 350, 50),
        ("Gaseosa Lata", 500, 320, 50),
        ("Tropical", 500, 320, 50),
        ("Gatorade", 1000, 700, 50),
        ("Agua", 1000, 600, 50),
        ("Cerveza Nacional", 1500, 1000, 50),
        ("Cerveza Importada", 1900, 1350, 50),
    ]
    extras = [
        ("Salchichas (Extra)", 700, 450, 999999),
        ("Papas (Extra)", 700, 450, 999999),
        ("Queso Amarillo", 200, 120, 999999),
        ("Jamón", 400, 250, 999999),
        ("Tocino", 700, 500, 999999),
        ("Helados Artesanales", 600, 350, 999999),
    ]

    for name, sale_price, cost_price, stock in comidas:
        conn.execute(
            """
            INSERT INTO products(name, category, sale_price, cost_price, stock, created_at)
            VALUES(?,?,?,?,?,?);
            """,
            (name, "comida", sale_price, cost_price, stock, now),
        )
    for name, sale_price, cost_price, stock in bebidas:
        conn.execute(
            """
            INSERT INTO products(name, category, sale_price, cost_price, stock, created_at)
            VALUES(?,?,?,?,?,?);
            """,
            (name, "bebida", sale_price, cost_price, stock, now),
        )
    for name, sale_price, cost_price, stock in extras:
        conn.execute(
            """
            INSERT INTO products(name, category, sale_price, cost_price, stock, created_at)
            VALUES(?,?,?,?,?,?);
            """,
            (name, "extra", sale_price, cost_price, stock, now),
        )


def get_setting(k: str, default: str = "") -> str:
    conn = get_conn()
    try:
        row = conn.execute("SELECT v FROM settings WHERE k=?;", (k,)).fetchone()
        return row["v"] if row else default
    finally:
        conn.close()


def set_setting(k: str, v: str):
    conn = get_conn()
    try:
        conn.execute(
            "INSERT INTO settings(k,v) VALUES(?,?) ON CONFLICT(k) DO UPDATE SET v=excluded.v;",
            (k, v),
        )
        conn.commit()
    finally:
        conn.close()


def list_products():
    conn = get_conn()
    try:
        return conn.execute(
            "SELECT id, name, category, sale_price, cost_price, stock, image_mime FROM products ORDER BY category, name;"
        ).fetchall()
    finally:
        conn.close()


def get_product(product_id: int):
    conn = get_conn()
    try:
        return conn.execute(
            "SELECT id, name, category, sale_price, cost_price, stock, image_blob, image_mime FROM products WHERE id=?;",
            (product_id,),
        ).fetchone()
    finally:
        conn.close()


def adjust_stock(product_id: int, delta: int):
    conn = get_conn()
    try:
        row = conn.execute("SELECT stock FROM products WHERE id=?;", (product_id,)).fetchone()
        if row is None:
            return False, "Producto no encontrado"
        new_stock = row["stock"] + delta
        if new_stock < 0:
            return False, "No hay stock suficiente"
        conn.execute("UPDATE products SET stock=? WHERE id=?;", (new_stock, product_id))
        conn.commit()
        return True, ""
    finally:
        conn.close()


# ============================================================
# "Cuentas" en memoria
# ============================================================
cuentas = {"Cuenta General": {"items": []}}
cuenta_actual = "Cuenta General"


def calcular_total(items):
    return sum(i["sale_price"] for i in items)


def calcular_costo(items):
    return sum(i["cost_price"] for i in items)


# ============================================================
# POS routes
# ============================================================
@app.route("/", methods=["GET"])
def index():
    prods = list_products()
    items = cuentas[cuenta_actual]["items"]
    total = calcular_total(items)
    return render_template_string(
        TEMPLATE_PRINCIPAL,
        cuentas=cuentas,
        cuenta_actual=cuenta_actual,
        products=prods,
        total=total,
    )


@app.route("/seleccionar_cuenta", methods=["POST"])
def seleccionar_cuenta():
    global cuenta_actual
    nombre = request.form.get("cuenta", "")
    if nombre in cuentas:
        cuenta_actual = nombre
    return redirect(url_for("index"))


@app.route("/nueva_cuenta", methods=["POST"])
def nueva_cuenta():
    global cuenta_actual
    nombre = request.form.get("nombre_cuenta", "").strip()
    if not nombre:
        flash("Debe escribir un nombre para la cuenta.", "error")
        return redirect(url_for("index"))
    if nombre in cuentas:
        flash("Ya existe una cuenta con ese nombre.", "error")
        return redirect(url_for("index"))
    cuentas[nombre] = {"items": []}
    cuenta_actual = nombre
    return redirect(url_for("index"))


@app.route("/eliminar_cuenta", methods=["POST"])
def eliminar_cuenta():
    global cuenta_actual
    nombre = request.form.get("cuenta_eliminar", "")

    if nombre == "Cuenta General":
        flash("La Cuenta General no se puede borrar.", "error")
        return redirect(url_for("index"))

    if nombre not in cuentas:
        return redirect(url_for("index"))

    if cuentas[nombre]["items"]:
        flash("Solo se puede borrar una cuenta vacía.", "error")
        return redirect(url_for("index"))

    del cuentas[nombre]
    if cuenta_actual == nombre:
        cuenta_actual = "Cuenta General"

    flash(f"Cuenta '{nombre}' eliminada.", "ok")
    return redirect(url_for("index"))


@app.route("/agregar_item", methods=["POST"])
def agregar_item():
    product_id = request.form.get("product_id", type=int)
    if not product_id:
        flash("Producto inválido.", "error")
        return redirect(url_for("index"))

    p = get_product(product_id)
    if not p:
        flash("Producto no encontrado.", "error")
        return redirect(url_for("index"))

    if p["stock"] <= 0:
        flash(f"No hay inventario disponible de {p['name']}.", "error")
        return redirect(url_for("index"))

    ok, msg = adjust_stock(product_id, -1)
    if not ok:
        flash(msg, "error")
        return redirect(url_for("index"))

    cuentas[cuenta_actual]["items"].append(
        {
            "product_id": p["id"],
            "name": p["name"],
            "category": p["category"],
            "sale_price": int(p["sale_price"]),
            "cost_price": int(p["cost_price"]),
            # control de despacho antes de cobrar (opcional)
            "delivered": 0,
        }
    )
    return redirect(url_for("index"))


@app.route("/eliminar_item", methods=["POST"])
def eliminar_item():
    idx = request.form.get("index", type=int)
    if idx is None:
        return redirect(url_for("index"))

    items = cuentas[cuenta_actual]["items"]
    if 0 <= idx < len(items):
        item = items.pop(idx)
        adjust_stock(item["product_id"], +1)
    return redirect(url_for("index"))


@app.route("/toggle_item_delivered", methods=["POST"])
def toggle_item_delivered():
    idx = request.form.get("index", type=int)
    if idx is None:
        return redirect(url_for("index"))
    items = cuentas[cuenta_actual]["items"]
    if 0 <= idx < len(items):
        items[idx]["delivered"] = 0 if items[idx].get("delivered", 0) else 1
    return redirect(url_for("index"))


@app.route("/cobrar", methods=["POST"])
def cobrar():
    global cuenta_actual

    items = cuentas[cuenta_actual]["items"]
    if not items:
        flash("La cuenta está vacía.", "error")
        return redirect(url_for("index"))

    regalia_code = (request.form.get("regalia_code") or "").strip()
    es_regalia = bool(regalia_code) and (regalia_code == get_setting("regalia_code", REGALIA_CODE_DEFAULT))

    payment_method = (request.form.get("payment_method") or "efectivo").strip().lower()
    if payment_method not in {"efectivo", "tarjeta", "sinpe"}:
        payment_method = "efectivo"

    total = calcular_total(items)
    cost_total = calcular_costo(items)
    profit = total - cost_total

    cash_received = (request.form.get("cash_received") or "").strip()
    cash_received_val = None
    change_given = None

    if es_regalia:
        cash_received_val = 0
        change_given = 0
        payment_method = payment_method or "efectivo"
    else:
        if payment_method == "efectivo":
            if not cash_received:
                flash("En efectivo debes escribir 'Paga con'.", "error")
                return redirect(url_for("index"))
            try:
                cash_received_val = int(float(cash_received))
            except ValueError:
                flash("Pago en efectivo inválido.", "error")
                return redirect(url_for("index"))
            if cash_received_val < total:
                flash(f"El pago (₡{cash_received_val}) no alcanza el total (₡{total}).", "error")
                return redirect(url_for("index"))
            change_given = cash_received_val - total
        else:
            cash_received_val = None
            change_given = 0

    conn = get_conn()
    try:
        now = datetime.now().isoformat()

        # status='paid' al cobrar; luego cocina lo “cierra” cuando entrega todo
        cur = conn.execute(
            """
            INSERT INTO sales(created_at, account_name, total, cost_total, profit, is_gift, cash_received, change_given, payment_method, status, closed_at)
            VALUES(?,?,?,?,?,?,?,?,?,?,?);
            """,
            (
                now,
                cuenta_actual,
                total,
                cost_total,
                profit,
                1 if es_regalia else 0,
                cash_received_val,
                change_given,
                payment_method,
                "paid",
                None,
            ),
        )
        sale_id = cur.lastrowid

        for it in items:
            delivered = 1 if it.get("delivered", 0) else 0
            delivered_at = now if delivered else None
            conn.execute(
                """
                INSERT INTO sale_items(sale_id, product_id, name, category, sale_price, cost_price, created_at, delivered, delivered_at)
                VALUES(?,?,?,?,?,?,?,?,?);
                """,
                (
                    sale_id,
                    it["product_id"],
                    it["name"],
                    it["category"],
                    it["sale_price"],
                    it["cost_price"],
                    now,
                    delivered,
                    delivered_at,
                ),
            )

        # si ya venía todo entregado, cerramos de una vez
        pending = conn.execute(
            "SELECT COUNT(*) AS c FROM sale_items WHERE sale_id=? AND delivered=0;",
            (sale_id,),
        ).fetchone()["c"]
        if pending == 0:
            conn.execute(
                "UPDATE sales SET status='closed', closed_at=? WHERE id=?;",
                (now, sale_id),
            )

        conn.commit()
    finally:
        conn.close()

    # Mensaje final
    if es_regalia:
        flash(f"✅ Cobrado como REGALÍA. Total real: ₡{total} · Cliente paga: ₡0", "ok")
    else:
        if payment_method == "efectivo":
            flash(
                f"✅ Cobrado en EFECTIVO. Total: ₡{total} · Paga con: ₡{cash_received_val} · Vuelto: ₡{change_given}",
                "ok",
            )
        elif payment_method == "tarjeta":
            flash(f"✅ Cobrado con TARJETA. Total: ₡{total}", "ok")
        else:
            flash(f"✅ Cobrado con SINPE. Total: ₡{total}", "ok")

    # Vaciar y si no es general, eliminar cuenta
    cuentas[cuenta_actual]["items"].clear()
    if cuenta_actual != "Cuenta General":
        del cuentas[cuenta_actual]
        cuenta_actual = "Cuenta General"

    return redirect(url_for("index"))


# ============================================================
# Cocina (control de despacho)
# ============================================================
def fetch_kitchen(limit: int = 60):
    """
    Trae ítems NO entregados de ventas pagadas (status != 'closed')
    """
    conn = get_conn()
    try:
        rows = conn.execute(
            """
            SELECT
              si.id AS sale_item_id,
              si.name,
              si.category,
              si.created_at AS item_created_at,
              s.id AS sale_id,
              s.account_name,
              s.created_at AS sale_created_at,
              s.status
            FROM sale_items si
            JOIN sales s ON s.id = si.sale_id
            WHERE si.delivered = 0
              AND s.status != 'closed'
            ORDER BY s.created_at ASC, si.id ASC
            LIMIT ?;
            """,
            (limit,),
        ).fetchall()
        return rows
    finally:
        conn.close()


def maybe_close_sale(conn: sqlite3.Connection, sale_id: int):
    pending = conn.execute(
        "SELECT COUNT(*) AS c FROM sale_items WHERE sale_id=? AND delivered=0;",
        (sale_id,),
    ).fetchone()["c"]
    if pending == 0:
        now = datetime.now().isoformat()
        conn.execute(
            "UPDATE sales SET status='closed', closed_at=? WHERE id=?;",
            (now, sale_id),
        )


@app.route("/cocina")
def cocina():
    rows = fetch_kitchen(limit=200)

    # Agrupar por sale_id
    grouped = {}
    for r in rows:
        sid = r["sale_id"]
        if sid not in grouped:
            grouped[sid] = {
                "sale_id": sid,
                "account_name": r["account_name"],
                "sale_created_at": r["sale_created_at"],
                "items": [],
            }
        grouped[sid]["items"].append(
            {
                "sale_item_id": r["sale_item_id"],
                "name": r["name"],
                "category": r["category"],
                "item_created_at": r["item_created_at"],
            }
        )

    # Orden en lista
    sales_list = sorted(grouped.values(), key=lambda x: x["sale_created_at"])
    return render_template_string(TEMPLATE_COCINA, sales=sales_list)


@app.route("/cocina/toggle", methods=["POST"])
def cocina_toggle():
    sale_item_id = request.form.get("sale_item_id", type=int)
    if not sale_item_id:
        return redirect(url_for("cocina"))

    conn = get_conn()
    try:
        row = conn.execute(
            "SELECT id, sale_id, delivered FROM sale_items WHERE id=?;",
            (sale_item_id,),
        ).fetchone()
        if not row:
            return redirect(url_for("cocina"))

        new_val = 0 if row["delivered"] else 1
        delivered_at = datetime.now().isoformat() if new_val == 1 else None

        conn.execute(
            "UPDATE sale_items SET delivered=?, delivered_at=? WHERE id=?;",
            (new_val, delivered_at, sale_item_id),
        )

        # si con esto ya no queda nada pendiente, cerramos la venta
        maybe_close_sale(conn, row["sale_id"])

        conn.commit()
    finally:
        conn.close()

    return redirect(url_for("cocina"))


# ============================================================
# Images
# ============================================================
@app.route("/img/<int:product_id>")
def product_image(product_id: int):
    p = get_product(product_id)
    if not p or not p["image_blob"]:
        # 1x1 gif transparente
        return Response(
            b"GIF89a\x01\x00\x01\x00\x80\x00\x00\x00\x00\x00\xff\xff\xff!\xf9\x04\x01\x00\x00\x00\x00,\x00\x00\x00\x00\x01\x00\x01\x00\x00\x02\x02D\x01\x00;",
            mimetype="image/gif",
        )
    return Response(p["image_blob"], mimetype=p["image_mime"] or "image/jpeg")


# ============================================================
# Ajustes / Inventario CRUD
# ============================================================
@app.route("/ajustes", methods=["GET", "POST"])
def ajustes():
    if request.method == "POST":
        action = (request.form.get("action") or "").strip()

        # cambiar clave regalía
        if action == "update_regalia":
            nueva = (request.form.get("regalia_code") or "").strip()
            if not nueva:
                flash("La clave de regalías no puede estar vacía.", "error")
            else:
                set_setting("regalia_code", nueva)
                flash("Clave de regalías actualizada.", "ok")
            return redirect(url_for("ajustes"))

        # agregar producto
        if action == "add_product":
            name = (request.form.get("name") or "").strip()
            category = (request.form.get("category") or "comida").strip().lower()
            sale_price = request.form.get("sale_price", type=int)
            cost_price = request.form.get("cost_price", type=int)
            stock = request.form.get("stock", type=int)

            if category not in {"comida", "bebida", "extra"}:
                category = "comida"

            if not name:
                flash("Nombre requerido.", "error")
                return redirect(url_for("ajustes"))

            if sale_price is None or cost_price is None or stock is None:
                flash("Precio venta, costo y stock deben ser números.", "error")
                return redirect(url_for("ajustes"))

            img = request.files.get("image")
            image_blob = None
            image_mime = None
            if img and img.filename:
                image_blob = img.read()
                image_mime = img.mimetype or "image/jpeg"

            conn = get_conn()
            try:
                conn.execute(
                    """
                    INSERT INTO products(name, category, sale_price, cost_price, stock, image_blob, image_mime, created_at)
                    VALUES(?,?,?,?,?,?,?,?);
                    """,
                    (
                        name,
                        category,
                        int(sale_price),
                        int(cost_price),
                        int(stock),
                        image_blob,
                        image_mime,
                        datetime.now().isoformat(),
                    ),
                )
                conn.commit()
                flash("Producto agregado.", "ok")
            except sqlite3.IntegrityError:
                flash("Ya existe un producto con ese nombre.", "error")
            finally:
                conn.close()

            return redirect(url_for("ajustes"))

        # editar producto
        if action == "update_product":
            product_id = request.form.get("product_id", type=int)
            if not product_id:
                return redirect(url_for("ajustes"))

            name = (request.form.get("name") or "").strip()
            category = (request.form.get("category") or "comida").strip().lower()
            sale_price = request.form.get("sale_price", type=int)
            cost_price = request.form.get("cost_price", type=int)
            stock = request.form.get("stock", type=int)

            if category not in {"comida", "bebida", "extra"}:
                category = "comida"

            if not name or sale_price is None or cost_price is None or stock is None:
                flash("Campos inválidos para actualizar.", "error")
                return redirect(url_for("ajustes"))

            img = request.files.get("image")
            set_img = (request.form.get("set_img") or "").strip()  # "1" para reemplazar

            conn = get_conn()
            try:
                if img and img.filename and set_img == "1":
                    conn.execute(
                        """
                        UPDATE products
                        SET name=?, category=?, sale_price=?, cost_price=?, stock=?, image_blob=?, image_mime=?
                        WHERE id=?;
                        """,
                        (
                            name,
                            category,
                            int(sale_price),
                            int(cost_price),
                            int(stock),
                            img.read(),
                            img.mimetype or "image/jpeg",
                            product_id,
                        ),
                    )
                else:
                    conn.execute(
                        """
                        UPDATE products
                        SET name=?, category=?, sale_price=?, cost_price=?, stock=?
                        WHERE id=?;
                        """,
                        (name, category, int(sale_price), int(cost_price), int(stock), product_id),
                    )
                conn.commit()
                flash("Producto actualizado.", "ok")
            except sqlite3.IntegrityError:
                flash("No se pudo actualizar: el nombre ya existe.", "error")
            finally:
                conn.close()

            return redirect(url_for("ajustes"))

        # borrar producto
        if action == "delete_product":
            product_id = request.form.get("product_id", type=int)
            if not product_id:
                return redirect(url_for("ajustes"))

            conn = get_conn()
            try:
                conn.execute("DELETE FROM products WHERE id=?;", (product_id,))
                conn.commit()
                flash("Producto eliminado.", "ok")
            finally:
                conn.close()

            return redirect(url_for("ajustes"))

    # GET
    prods = list_products()
    regalia_mask = "********"
    return render_template_string(
        TEMPLATE_AJUSTES,
        regalia_code=regalia_mask,
        products=prods,
    )


# ============================================================
# Reportes + Export
# ============================================================
def period_bounds(periodo: str):
    today = date.today()
    if periodo == "hoy":
        start = today
        end = today + timedelta(days=1)
    elif periodo == "semana":
        start = today - timedelta(days=7)
        end = today + timedelta(days=1)
    elif periodo == "mes":
        start = today.replace(day=1)
        if start.month == 12:
            end = date(start.year + 1, 1, 1)
        else:
            end = date(start.year, start.month + 1, 1)
    else:
        start = date(1970, 1, 1)
        end = date(2999, 1, 1)
    return start, end


def fetch_report(periodo: str):
    start, end = period_bounds(periodo)
    start_dt = datetime.combine(start, datetime.min.time()).isoformat()
    end_dt = datetime.combine(end, datetime.min.time()).isoformat()

    conn = get_conn()
    try:
        rows = conn.execute(
            """
            SELECT
              id, created_at, account_name, total, cost_total, profit,
              is_gift, cash_received, change_given, payment_method,
              status, closed_at
            FROM sales
            WHERE created_at >= ? AND created_at < ?
            ORDER BY created_at DESC;
            """,
            (start_dt, end_dt),
        ).fetchall()

        totals = conn.execute(
            """
            SELECT
              COALESCE(SUM(CASE WHEN is_gift=0 THEN total END),0) AS total_vendido,
              COALESCE(SUM(CASE WHEN is_gift=1 THEN total END),0) AS total_regalias,
              COALESCE(SUM(cost_total),0) AS costo_total,
              COALESCE(SUM(profit),0) AS ganancia_total,
              COUNT(*) AS n_ventas
            FROM sales
            WHERE created_at >= ? AND created_at < ?;
            """,
            (start_dt, end_dt),
        ).fetchone()

        daily = conn.execute(
            """
            SELECT
              substr(created_at,1,10) AS dia,
              COUNT(*) AS n,
              COALESCE(SUM(CASE WHEN is_gift=0 THEN total END),0) AS vendido,
              COALESCE(SUM(cost_total),0) AS costo,
              COALESCE(SUM(profit),0) AS ganancia
            FROM sales
            WHERE created_at >= ? AND created_at < ?
            GROUP BY substr(created_at,1,10)
            ORDER BY dia DESC;
            """,
            (start_dt, end_dt),
        ).fetchall()

        weekday = conn.execute(
            """
            SELECT
              strftime('%w', created_at) AS weekday_num,
              COUNT(*) AS n,
              COALESCE(SUM(CASE WHEN is_gift=0 THEN total END),0) AS vendido,
              COALESCE(SUM(cost_total),0) AS costo,
              COALESCE(SUM(profit),0) AS ganancia
            FROM sales
            WHERE created_at >= ? AND created_at < ?
            GROUP BY strftime('%w', created_at)
            ORDER BY weekday_num;
            """,
            (start_dt, end_dt),
        ).fetchall()

        return rows, totals, daily, weekday
    finally:
        conn.close()


@app.route("/reportes")
def reportes():
    periodo = request.args.get("periodo", "todo")
    sales, totals, daily, weekday = fetch_report(periodo)

    day_names = ["Domingo", "Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado"]
    weekday_pretty = [
        {
            "dia": day_names[int(r["weekday_num"])],
            "n": r["n"],
            "vendido": r["vendido"],
            "costo": r["costo"],
            "ganancia": r["ganancia"],
        }
        for r in weekday
    ]

    return render_template_string(
        TEMPLATE_REPORTES,
        periodo=periodo,
        sales=sales,
        totals=totals,
        daily=daily,
        weekday=weekday_pretty,
    )


@app.route("/exportar")
def exportar():
    periodo = request.args.get("periodo", "todo")
    sales, totals, daily, weekday = fetch_report(periodo)

    # Excel si está openpyxl, si no CSV
    try:
        from openpyxl import Workbook
    except Exception:
        output = io.StringIO()
        output.write("id,fecha,cuenta,total,costo,ganancia,regalia,metodo,paga_con,vuelto,estado,closed_at\n")
        for s in sales:
            output.write(
                f'{s["id"]},{s["created_at"]},{s["account_name"]},{s["total"]},{s["cost_total"]},{s["profit"]},{s["is_gift"]},{s["payment_method"]},{s["cash_received"]},{s["change_given"]},{s["status"]},{s["closed_at"]}\n'
            )
        mem = io.BytesIO(output.getvalue().encode("utf-8"))
        return send_file(mem, mimetype="text/csv", as_attachment=True, download_name=f"historial_{periodo}.csv")

    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"

    ws.append(["ID", "Fecha", "Cuenta", "Total", "Costo", "Ganancia", "Regalía", "Método", "Paga con", "Vuelto", "Estado", "Cerrada en"])
    for s in sales:
        ws.append(
            [
                s["id"],
                s["created_at"],
                s["account_name"],
                s["total"],
                s["cost_total"],
                s["profit"],
                "Sí" if s["is_gift"] else "No",
                s["payment_method"] or "",
                s["cash_received"] if s["cash_received"] is not None else "",
                s["change_given"] if s["change_given"] is not None else "",
                s["status"] or "",
                s["closed_at"] or "",
            ]
        )

    ws2 = wb.create_sheet("Totales")
    ws2.append(["Periodo", periodo])
    ws2.append(["Número de ventas", totals["n_ventas"]])
    ws2.append(["Total vendido", totals["total_vendido"]])
    ws2.append(["Total regalías (valor)", totals["total_regalias"]])
    ws2.append(["Costo total", totals["costo_total"]])
    ws2.append(["Ganancia total", totals["ganancia_total"]])

    ws3 = wb.create_sheet("Ventas por día")
    ws3.append(["Día", "Ventas", "Vendido", "Costo", "Ganancia"])
    for d in daily:
        ws3.append([d["dia"], d["n"], d["vendido"], d["costo"], d["ganancia"]])

    ws4 = wb.create_sheet("Por día semana")
    ws4.append(["Día semana", "Ventas", "Vendido", "Costo", "Ganancia"])
    for w in weekday:
        ws4.append([w["dia"], w["n"], w["vendido"], w["costo"], w["ganancia"]])

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    return send_file(
        out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"historial_{periodo}.xlsx",
    )


# ============================================================
# Templates
# ============================================================
 TEMPLATE_PRINCIPAL = """
<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <title>Sport Spot | POS Pro</title>
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <script src="https://cdn.tailwindcss.com"></script>
  <script src="https://unpkg.com/lucide@latest"></script>
  <style>
    @import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@300;400;600;800&display=swap');
    body { 
      font-family: 'Plus Jakarta Sans', sans-serif;
      background: radial-gradient(circle at top left, #166534 0%, #064e3b 40%, #020617 100%);
      min-height: 100vh;
    }
    .glass {
      background: rgba(255, 255, 255, 0.03);
      backdrop-filter: blur(12px);
      border: 1px solid rgba(255, 255, 255, 0.1);
    }
    .custom-scrollbar::-webkit-scrollbar { width: 6px; }
    .custom-scrollbar::-webkit-scrollbar-track { background: transparent; }
    .custom-scrollbar::-webkit-scrollbar-thumb { background: rgba(255,255,255,0.1); border-radius: 10px; }
  </style>
</head>
<body class="text-slate-200 p-4 lg:p-6">


<div class="max-w-[1600px] mx-auto">
  <header class="glass rounded-3xl p-5 mb-6 flex flex-col md:flex-row justify-between items-center gap-4 shadow-2xl">
    <div class="flex items-center gap-4">
      <div class="bg-green-500 p-3 rounded-2xl shadow-lg shadow-green-500/20">
        <i data-lucide="layout-dashboard" class="text-white w-6 h-6"></i>
      </div>
      <div>
        <h1 class="text-xl font-extrabold tracking-tight text-white">SPORT SPOT <span class="text-green-400">POS</span></h1>
        <p class="text-xs text-slate-400 font-medium uppercase tracking-widest">Sistema de Gestión e Inventarios</p>
      </div>
    </div>
    <nav class="flex gap-2">
      <a href="{{ url_for('cocina') }}" class="flex items-center gap-2 px-4 py-2 rounded-xl glass hover:bg-white/10 transition-all text-sm font-semibold">
        <i data-lucide="utensils" class="w-4 h-4 text-orange-400"></i> Cocina
      </a>
      <a href="{{ url_for('reportes') }}" class="flex items-center gap-2 px-4 py-2 rounded-xl glass hover:bg-white/10 transition-all text-sm font-semibold">
        <i data-lucide="bar-chart-3" class="w-4 h-4 text-blue-400"></i> Reportes
      </a>
      <a href="{{ url_for('ajustes') }}" class="flex items-center gap-2 px-4 py-2 rounded-xl glass hover:bg-white/10 transition-all text-sm font-semibold">
        <i data-lucide="settings" class="w-4 h-4 text-slate-400"></i> Ajustes
      </a>
    </nav>
  </header>


  <div class="grid grid-cols-1 lg:grid-cols-12 gap-6">
    
    <main class="lg:col-span-8">
      <div class="flex items-center justify-between mb-6 px-2">
        <h2 class="text-lg font-bold flex items-center gap-2">
          <i data-lucide="shopping-bag" class="text-green-400"></i> Catálogo de Productos
        </h2>
        <div class="text-xs bg-white/5 px-3 py-1 rounded-full border border-white/10 text-slate-400">
          {{ products|length }} productos disponibles
        </div>
      </div>


      <div class="grid grid-cols-2 md:grid-cols-3 xl:grid-cols-4 gap-4 custom-scrollbar overflow-y-auto max-h-[75vh] pr-2">
        {% for p in products %}
        {% set sin_stock = p['stock'] <= 0 %}
        <div class="glass rounded-2xl p-3 group hover:border-green-500/50 transition-all duration-300 {% if sin_stock %}opacity-50 grayscale{% endif %}">
          <div class="relative h-32 rounded-xl overflow-hidden mb-3 shadow-inner bg-black/20">
            <img src="{{ url_for('product_image', product_id=p['id']) }}" class="w-full h-full object-cover group-hover:scale-110 transition-transform duration-500">
            <div class="absolute top-2 right-2 bg-black/60 backdrop-blur-md px-2 py-1 rounded-lg text-[10px] font-bold border border-white/10">
              STOCK: {{ p['stock'] }}
            </div>
          </div>
          <div class="flex justify-between items-start mb-2">
            <div>
              <h3 class="font-bold text-sm leading-tight mb-1">{{ p['name'] }}</h3>
              <span class="text-[10px] uppercase font-bold text-green-400 tracking-tighter">{{ p['category'] }}</span>
            </div>
            <span class="font-extrabold text-sm text-white">₡{{ p['sale_price'] }}</span>
          </div>
          <form method="post" action="{{ url_for('agregar_item') }}">
            <input type="hidden" name="product_id" value="{{ p['id'] }}">
            <button type="submit" {% if sin_stock %}disabled{% endif %} 
              class="w-full py-2 rounded-xl bg-green-500 hover:bg-green-400 text-green-950 font-bold text-xs transition-colors shadow-lg shadow-green-500/20 active:scale-95 disabled:bg-slate-700 disabled:text-slate-500">
              AGREGAR
            </button>
          </form>
        </div>
        {% endfor %}
      </div>
    </main>


    <aside class="lg:col-span-4 space-y-6">
      <div class="glass rounded-3xl p-6 shadow-2xl sticky top-6">
        <h2 class="text-lg font-bold mb-4 flex items-center gap-2">
          <i data-lucide="receipt" class="text-blue-400"></i> Cuenta Actual
        </h2>


        <div class="flex flex-wrap gap-2 mb-4">
          {% for nombre, data in cuentas.items() %}
          <form method="post" action="{{ url_for('seleccionar_cuenta') }}">
            <input type="hidden" name="cuenta" value="{{ nombre }}">
            <button type="submit" class="px-3 py-1.5 rounded-xl text-xs font-bold transition-all
              {% if nombre == cuenta_actual %}
                bg-blue-500 text-white shadow-lg shadow-blue-500/30
              {% else %}
                glass hover:bg-white/10 text-slate-400
              {% endif %}">
              {{ nombre }} <span class="opacity-60 ml-1">{{ data['items']|length }}</span>
            </button>
          </form>
          {% endfor %}
        </div>


        <div class="flex gap-2 mb-6">
          <form class="flex-1 flex gap-2" method="post" action="{{ url_for('nueva_cuenta') }}">
            <input class="flex-1 bg-black/20 border border-white/10 rounded-xl px-3 py-2 text-xs focus:border-blue-500 outline-none transition-all" 
              type="text" name="nombre_cuenta" placeholder="Nueva mesa/cuenta...">
            <button class="bg-white/10 hover:bg-white/20 p-2 rounded-xl transition-all" type="submit">
              <i data-lucide="plus" class="w-4 h-4"></i>
            </button>
          </form>
          <form method="post" action="{{ url_for('eliminar_cuenta') }}">
            <input type="hidden" name="cuenta_eliminar" value="{{ cuenta_actual }}">
            <button class="p-2 rounded-xl bg-rose-500/10 hover:bg-rose-500 text-rose-500 hover:text-white transition-all border border-rose-500/20" type="submit">
              <i data-lucide="trash-2" class="w-4 h-4"></i>
            </button>
          </form>
        </div>


        <div class="bg-black/20 rounded-2xl p-2 mb-6 min-h-[150px] max-h-[300px] overflow-y-auto custom-scrollbar">
          {% if cuentas[cuenta_actual]['items'] %}
            {% for item in cuentas[cuenta_actual]['items'] %}
            <div class="flex justify-between items-center p-3 rounded-xl hover:bg-white/5 transition-all group">
              <div class="flex-1">
                <div class="flex items-center gap-2">
                  <span class="text-[9px] bg-white/10 px-1.5 py-0.5 rounded text-slate-400 font-bold uppercase">{{ item.category }}</span>
                  <span class="text-sm font-semibold">{{ item.name }}</span>
                </div>
                <form method="post" action="{{ url_for('toggle_item_delivered') }}" class="mt-1">
                  <input type="hidden" name="index" value="{{ loop.index0 }}">
                  <button type="submit" class="text-[10px] font-bold flex items-center gap-1 {% if item.delivered %}text-green-400{% else %}text-amber-500{% endif %}">
                    <i data-lucide="{% if item.delivered %}check-circle{% else %}clock{% endif %}" class="w-3 h-3"></i>
                    {% if item.delivered %}Entregado{% else %}Pendiente{% endif %}
                  </button>
                </form>
              </div>
              <div class="flex items-center gap-3">
                <span class="text-sm font-bold text-white">₡{{ item.sale_price }}</span>
                <form method="post" action="{{ url_for('eliminar_item') }}">
                  <input type="hidden" name="index" value="{{ loop.index0 }}">
                  <button type="submit" class="text-slate-500 hover:text-rose-500 transition-colors">
                    <i data-lucide="x-circle" class="w-4 h-4"></i>
                  </button>
                </form>
              </div>
            </div>
            {% endfor %}
          {% else %}
            <div class="h-full flex flex-col items-center justify-center text-slate-500 py-10">
              <i data-lucide="shopping-cart" class="w-8 h-8 mb-2 opacity-20"></i>
              <p class="text-xs font-medium italic">Cuenta vacía</p>
            </div>
          {% endif %}
        </div>


        <div class="border-t border-white/10 pt-4 space-y-4">
          <div class="flex justify-between items-end">
            <span class="text-xs font-bold text-slate-400 uppercase tracking-widest">Total a pagar</span>
            <span class="text-3xl font-black text-white leading-none">₡{{ total }}</span>
          </div>


          <form method="post" action="{{ url_for('cobrar') }}" class="space-y-3">
            <div class="grid grid-cols-2 gap-2">
              <div class="space-y-1">
                <label class="text-[10px] font-bold text-slate-500 uppercase ml-2">Pago</label>
                <select name="payment_method" class="w-full bg-black/20 border border-white/10 rounded-xl px-3 py-2 text-xs text-white outline-none focus:border-green-500">
                  <option value="efectivo">Efectivo</option>
                  <option value="tarjeta">Tarjeta</option>
                  <option value="sinpe">SINPE</option>
                </select>
              </div>
              <div class="space-y-1">
                <label class="text-[10px] font-bold text-slate-500 uppercase ml-2">Paga con</label>
                <input name="cash_received" class="w-full bg-black/20 border border-white/10 rounded-xl px-3 py-2 text-xs text-white outline-none focus:border-green-500" placeholder="Ej: 5000">
              </div>
            </div>
            <div class="space-y-1">
              <label class="text-[10px] font-bold text-slate-500 uppercase ml-2">Código Regalía</label>
              <input type="password" name="regalia_code" class="w-full bg-black/20 border border-white/10 rounded-xl px-3 py-2 text-xs text-white outline-none focus:border-yellow-500" placeholder="••••">
            </div>
            <button type="submit" class="w-full py-4 bg-green-500 hover:bg-green-400 text-green-950 font-black rounded-2xl shadow-xl shadow-green-500/20 transition-all active:scale-[0.98] flex items-center justify-center gap-2">
              <i data-lucide="banknote" class="w-5 h-5"></i> FINALIZAR VENTA
            </button>
          </form>
        </div>
      </div>
    </aside>
  </div>


  {% with messages = get_flashed_messages(with_categories=true) %}
    {% if messages %}
    <div class="fixed bottom-6 left-1/2 -translate-x-1/2 z-50 flex flex-col gap-2 w-full max-w-md px-4">
      {% for category, msg in messages %}
      <div class="glass border-l-4 {% if category == 'ok' %}border-green-500 bg-green-500/10{% else %}border-rose-500 bg-rose-500/10{% endif %} p-4 rounded-2xl shadow-2xl flex items-center justify-between animate-bounce">
        <p class="text-sm font-bold {% if category == 'ok' %}text-green-400{% else %}text-rose-400{% endif %}">{{ msg }}</p>
        <button onclick="this.parentElement.remove()" class="text-white/50 hover:text-white">
          <i data-lucide="x" class="w-4 h-4"></i>
        </button>
      </div>
      {% endfor %}
    </div>
    {% endif %}
  {% endwith %}


</div>


<script>
  // Inicializar Iconos Lucide
  lucide.createIcons();


  // Auto-cerrar alertas flash después de 4 segundos
  setTimeout(() => {
    document.querySelectorAll('.animate-bounce').forEach(el => el.remove());
  }, 4000);
</script>


</body>
</html>
""”



TEMPLATE_AJUSTES = """
<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <title>Ajustes · Sport Spot</title>
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <style>
    * { box-sizing:border-box; font-family: system-ui, -apple-system, Segoe UI, sans-serif; }
    body { margin:0; background: radial-gradient(circle at top left, #4caf50 0, #1b5e20 40%, #0f2d1f 100%); color:#f5f5f5; }
    .app { max-width: 1200px; margin:0 auto; padding:16px; }
    a { color:#c8e6c9; }
    .card { background: rgba(3,18,10,0.9); border-radius:18px; padding:16px; border:1px solid rgba(129,199,132,0.35); box-shadow: 0 12px 30px rgba(0,0,0,0.4); margin: 14px 0; }
    h1 { margin: 0 0 10px; font-size: 1.2rem; text-transform: uppercase; letter-spacing: .08em; }
    h2 { margin:0 0 10px; font-size:1rem; }
    .row { display:flex; gap:10px; flex-wrap: wrap; align-items: center; }
    label { font-size:.85rem; opacity:.9; }
    input, select { padding:7px 12px; border-radius:999px; border:1px solid rgba(200,230,201,.35); background: rgba(0,0,0,.35); color:#f1f8e9; font-size: .9rem; }
    .btn { border:none; border-radius:999px; padding:7px 14px; font-size:.85rem; cursor:pointer; white-space:nowrap;}
    .btn-primary { background: linear-gradient(135deg, #cddc39, #8bc34a); color:#1b5e20; font-weight:800; }
    .btn-outline { background: transparent; color:#c8e6c9; border:1px solid rgba(200,230,201,.35); }
    .btn-danger { background: linear-gradient(135deg, #d32f2f, #b71c1c); color:#fff; font-weight:800; }
    table { width:100%; border-collapse: collapse; font-size:.85rem; background: rgba(0,0,0,.25); border-radius: 12px; overflow:hidden; }
    th, td { padding: 8px 10px; border-bottom: 1px solid rgba(255,255,255,0.06); vertical-align: middle; }
    th { text-align:left; background: rgba(0,0,0,.4); }
    .img { width:70px; height:44px; border-radius: 10px; overflow:hidden; background: rgba(0,0,0,.35); border:1px solid rgba(255,255,255,.06); }
    .img img{ width:100%; height:100%; object-fit: cover; }
    .flash { margin: 10px 0; padding: 10px 14px; border-radius: 999px; background: rgba(0,0,0,0.82); border:1px solid rgba(200,230,201,.5); }
    .flash.ok{ border-color: rgba(129,199,132,.8); color:#e8f5e9; }
    .flash.error{ border-color:#ef9a9a; color:#ffebee; }
    small { opacity:.8; }
  </style>
</head>
<body>
<div class="app">
  <a href="{{ url_for('index') }}">&larr; Volver</a>
  <h1>Ajustes</h1>

  {% with messages = get_flashed_messages(with_categories=true) %}
    {% if messages %}
      {% for category, msg in messages %}
        <div class="flash {{ category }}">{{ msg }}</div>
      {% endfor %}
    {% endif %}
  {% endwith %}

  <div class="card">
    <h2>Clave de regalías</h2>
    <form method="post">
      <input type="hidden" name="action" value="update_regalia">
      <div class="row">
        <div><small>Actual (oculta):</small> <code>{{ regalia_code }}</code></div>
      </div>
      <div class="row" style="margin-top:8px;">
        <label>Nueva clave:</label>
        <input type="password" name="regalia_code" placeholder="Escriba nueva clave">
        <button class="btn btn-primary" type="submit">Guardar</button>
      </div>
    </form>
  </div>

  <div class="card">
    <h2>Agregar producto</h2>
    <form method="post" enctype="multipart/form-data">
      <input type="hidden" name="action" value="add_product">
      <div class="row">
        <label>Nombre</label>
        <input name="name" placeholder="Ej: Nachos">
        <label>Categoría</label>
        <select name="category">
          <option value="comida">comida</option>
          <option value="bebida">bebida</option>
          <option value="extra">extra</option>
        </select>
      </div>
      <div class="row" style="margin-top:8px;">
        <label>Precio venta</label>
        <input name="sale_price" type="number" placeholder="₡">
        <label>Costo</label>
        <input name="cost_price" type="number" placeholder="₡">
        <label>Stock</label>
        <input name="stock" type="number" placeholder="Ej: 30">
      </div>
      <div class="row" style="margin-top:8px;">
        <label>Imagen (archivo)</label>
        <input name="image" type="file" accept="image/*">
        <button class="btn btn-primary" type="submit">Agregar</button>
      </div>
      <small>Si no subes imagen, se usa un placeholder.</small>
    </form>
  </div>

  <div class="card">
    <h2>Inventario (editar)</h2>
    <table>
      <thead>
        <tr>
          <th>Foto</th>
          <th>Nombre</th>
          <th>Categoría</th>
          <th>Venta</th>
          <th>Costo</th>
          <th>Stock</th>
          <th>Actualizar</th>
          <th>Borrar</th>
        </tr>
      </thead>
      <tbody>
        {% for p in products %}
        <tr>
          <td class="img"><img src="{{ url_for('product_image', product_id=p['id']) }}" alt=""></td>
          <td>
            <form method="post" enctype="multipart/form-data" style="display:grid; gap:6px;">
              <input type="hidden" name="action" value="update_product">
              <input type="hidden" name="product_id" value="{{ p['id'] }}">
              <input name="name" value="{{ p['name'] }}">
          </td>
          <td>
              <select name="category">
                <option value="comida" {% if p['category']=='comida' %}selected{% endif %}>comida</option>
                <option value="bebida" {% if p['category']=='bebida' %}selected{% endif %}>bebida</option>
                <option value="extra" {% if p['category']=='extra' %}selected{% endif %}>extra</option>
              </select>
          </td>
          <td><input name="sale_price" type="number" value="{{ p['sale_price'] }}"></td>
          <td><input name="cost_price" type="number" value="{{ p['cost_price'] }}"></td>
          <td><input name="stock" type="number" value="{{ p['stock'] }}"></td>
          <td>
              <div class="row">
                <label><small>Reemplazar imagen</small></label>
                <input type="hidden" name="set_img" value="0">
                <input type="checkbox" onclick="this.form.set_img.value = this.checked ? '1' : '0'">
                <input name="image" type="file" accept="image/*">
                <button class="btn btn-primary" type="submit">Guardar</button>
              </div>
            </form>
          </td>
          <td>
            <form method="post" onsubmit="return confirm('¿Eliminar producto?')">
              <input type="hidden" name="action" value="delete_product">
              <input type="hidden" name="product_id" value="{{ p['id'] }}">
              <button class="btn btn-danger" type="submit">Eliminar</button>
            </form>
          </td>
        </tr>
        {% endfor %}
      </tbody>
    </table>
    <small>“Costo” es lo que te costó a ti, “Venta” es el precio al cliente. La ganancia se calcula en reportes.</small>
  </div>
</div>
</body>
</html>
"""


TEMPLATE_REPORTES = """
<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <title>Reportes · Sport Spot</title>
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <style>
    * { box-sizing:border-box; font-family: system-ui, -apple-system, Segoe UI, sans-serif; }
    body { margin:0; background: radial-gradient(circle at top left, #4caf50 0, #1b5e20 40%, #0f2d1f 100%); color:#f5f5f5; }
    .app { max-width: 1200px; margin:0 auto; padding:16px; }
    a { color:#c8e6c9; }
    h1 { margin: 0 0 10px; font-size: 1.2rem; text-transform: uppercase; letter-spacing:.08em; }
    .bar { display:flex; flex-wrap:wrap; gap:8px; align-items:center; margin: 10px 0 14px; }
    .badge { display:inline-block; padding: 6px 12px; border-radius: 999px; border:1px solid rgba(200,230,201,.5); background: rgba(0,0,0,.35); font-size:.85rem; text-decoration:none; color:#e8f5e9; }
    .card { background: rgba(3,18,10,0.9); border-radius:18px; padding:14px; border:1px solid rgba(129,199,132,0.35); box-shadow: 0 12px 30px rgba(0,0,0,0.4); margin: 12px 0; }
    .summary { display:grid; grid-template-columns: repeat(auto-fit, minmax(220px, 1fr)); gap:10px; }
    .box { background: rgba(0,0,0,.25); border-radius: 12px; padding: 10px; font-size:.9rem; }
    table { width:100%; border-collapse: collapse; font-size:.85rem; background: rgba(0,0,0,.25); border-radius: 12px; overflow:hidden; }
    th, td { padding: 8px 10px; border-bottom: 1px solid rgba(255,255,255,0.06); }
    th { text-align:left; background: rgba(0,0,0,.4); }
    .chip { padding: 3px 10px; border-radius: 999px; font-size:.75rem; display:inline-block; }
    .c-gift { background: rgba(255,235,59,0.15); border:1px solid rgba(255,235,59,0.6); }
    .c-norm { background: rgba(129,199,132,0.15); border:1px solid rgba(129,199,132,0.6); }
    .c-open { background: rgba(255,255,255,0.10); border:1px solid rgba(255,255,255,0.25); }
    .c-closed { background: rgba(200,230,201,0.12); border:1px solid rgba(200,230,201,0.35); }
    .muted { opacity:.8; }
  </style>
</head>
<body>
<div class="app">
  <a href="{{ url_for('index') }}">&larr; Volver</a>
  <h1>Reportes</h1>

  <div class="bar">
    <span class="muted">Periodo:</span>
    <a class="badge" href="{{ url_for('reportes', periodo='hoy') }}">Hoy</a>
    <a class="badge" href="{{ url_for('reportes', periodo='semana') }}">Últimos 7 días</a>
    <a class="badge" href="{{ url_for('reportes', periodo='mes') }}">Mes</a>
    <a class="badge" href="{{ url_for('reportes', periodo='todo') }}">Todo</a>
    <span style="flex:1"></span>
    <a class="badge" href="{{ url_for('exportar', periodo=periodo) }}">⬇ Exportar historial</a>
  </div>

  <div class="card">
    <div class="summary">
      <div class="box"><strong>Número de ventas</strong><br>{{ totals['n_ventas'] }}</div>
      <div class="box"><strong>Total vendido</strong><br>₡{{ totals['total_vendido'] }}</div>
      <div class="box"><strong>Costo total</strong><br>₡{{ totals['costo_total'] }}</div>
      <div class="box"><strong>Ganancia total</strong><br>₡{{ totals['ganancia_total'] }}</div>
      <div class="box"><strong>Valor regalías</strong><br>₡{{ totals['total_regalias'] }}</div>
    </div>
  </div>

  <div class="card">
    <h2 style="margin:0 0 10px;">Ventas por día</h2>
    {% if daily %}
      <table>
        <thead>
          <tr><th>Día</th><th>#</th><th>Vendido</th><th>Costo</th><th>Ganancia</th></tr>
        </thead>
        <tbody>
          {% for d in daily %}
            <tr>
              <td>{{ d['dia'] }}</td>
              <td>{{ d['n'] }}</td>
              <td>₡{{ d['vendido'] }}</td>
              <td>₡{{ d['costo'] }}</td>
              <td>₡{{ d['ganancia'] }}</td>
            </tr>
          {% endfor %}
        </tbody>
      </table>
    {% else %}
      <div class="muted">No hay datos para este periodo.</div>
    {% endif %}
  </div>

  <div class="card">
    <h2 style="margin:0 0 10px;">Por día de semana</h2>
    {% if weekday %}
      <table>
        <thead>
          <tr><th>Día</th><th>#</th><th>Vendido</th><th>Costo</th><th>Ganancia</th></tr>
        </thead>
        <tbody>
          {% for w in weekday %}
            <tr>
              <td>{{ w['dia'] }}</td>
              <td>{{ w['n'] }}</td>
              <td>₡{{ w['vendido'] }}</td>
              <td>₡{{ w['costo'] }}</td>
              <td>₡{{ w['ganancia'] }}</td>
            </tr>
          {% endfor %}
        </tbody>
      </table>
    {% else %}
      <div class="muted">No hay datos para este periodo.</div>
    {% endif %}
  </div>

  <div class="card">
    <h2 style="margin:0 0 10px;">Historial de ventas</h2>
    {% if sales %}
      <table>
        <thead>
          <tr>
            <th>Fecha</th>
            <th>Cuenta</th>
            <th>Tipo</th>
            <th>Método</th>
            <th>Total</th>
            <th>Costo</th>
            <th>Ganancia</th>
            <th>Paga con</th>
            <th>Vuelto</th>
            <th>Estado</th>
          </tr>
        </thead>
        <tbody>
          {% for s in sales %}
            <tr>
              <td>{{ s['created_at'][:16].replace('T',' ') }}</td>
              <td>{{ s['account_name'] }}</td>
              <td>
                {% if s['is_gift'] %}
                  <span class="chip c-gift">Regalía</span>
                {% else %}
                  <span class="chip c-norm">Normal</span>
                {% endif %}
              </td>
              <td>{{ (s['payment_method'] or '')|upper }}</td>
              <td>₡{{ s['total'] }}</td>
              <td>₡{{ s['cost_total'] }}</td>
              <td>₡{{ s['profit'] }}</td>
              <td>{{ s['cash_received'] if s['cash_received'] is not none else '' }}</td>
              <td>{{ s['change_given'] if s['change_given'] is not none else '' }}</td>
              <td>
                {% if (s['status'] or '') == 'closed' %}
                  <span class="chip c-closed">Cerrada</span>
                {% else %}
                  <span class="chip c-open">En despacho</span>
                {% endif %}
              </td>
            </tr>
          {% endfor %}
        </tbody>
      </table>
    {% else %}
      <div class="muted">No hay ventas registradas.</div>
    {% endif %}
  </div>

</div>
</body>
</html>
"""


TEMPLATE_COCINA = """
<!doctype html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <title>Cocina · Sport Spot</title>
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <style>
    * { box-sizing:border-box; font-family: system-ui, -apple-system, Segoe UI, sans-serif; }
    body { margin:0; background: radial-gradient(circle at top left, #4caf50 0, #1b5e20 40%, #0f2d1f 100%); color:#f5f5f5; }
    .app { max-width: 1200px; margin:0 auto; padding:16px; }
    a { color:#c8e6c9; }
    h1 { margin: 0 0 10px; font-size: 1.2rem; text-transform: uppercase; letter-spacing:.08em; }
    .grid { display:grid; grid-template-columns: repeat(auto-fit, minmax(260px, 1fr)); gap: 12px; }
    .card { background: rgba(3,18,10,0.9); border-radius:18px; padding:14px; border:1px solid rgba(129,199,132,0.35); box-shadow: 0 12px 30px rgba(0,0,0,0.4); }
    .head { display:flex; justify-content:space-between; gap:10px; align-items:baseline; margin-bottom: 10px; }
    .sale { font-weight: 900; }
    .muted { opacity:.85; font-size:.85rem; }
    .item { background: rgba(0,0,0,.25); border:1px solid rgba(255,255,255,0.06); border-radius: 12px; padding: 10px; margin: 8px 0; display:flex; justify-content:space-between; gap: 10px; }
    .tag { font-size:.72rem; text-transform: uppercase; letter-spacing:.08em; padding: 2px 8px; border-radius: 999px; background: rgba(205,220,57,.12); color:#e6ee9c; display:inline-block; margin-right: 6px;}
    .btn { border:none; border-radius:999px; padding: 7px 12px; font-size:.82rem; cursor:pointer; white-space:nowrap;}
    .btn-primary { background: linear-gradient(135deg, #cddc39, #8bc34a); color:#1b5e20; font-weight:800;}
    .empty { opacity:.85; background: rgba(0,0,0,.25); border:1px solid rgba(255,255,255,0.06); border-radius: 14px; padding: 14px; }
  </style>
</head>
<body>
<div class="app">
  <a href="{{ url_for('index') }}">&larr; Volver al POS</a>
  <h1>Cocina (pendientes de entregar)</h1>

  {% if sales %}
    <div class="grid">
      {% for s in sales %}
        <div class="card">
          <div class="head">
            <div>
              <div class="sale">Venta #{{ s.sale_id }} · {{ s.account_name }}</div>
              <div class="muted">Creada: {{ s.sale_created_at[:16].replace('T',' ') }}</div>
            </div>
            <div class="muted">Pendientes: {{ s['items']|length }}</div>
          </div>

          {% for it in s['items'] %}
            <div class="item">
              <div>
                <span class="tag">{{ it.category }}</span>
                <strong>{{ it.name }}</strong>
                <div class="muted">Ítem #{{ it.sale_item_id }}</div>
              </div>
              <form method="post" action="{{ url_for('cocina_toggle') }}">
                <input type="hidden" name="sale_item_id" value="{{ it.sale_item_id }}">
                <button class="btn btn-primary" type="submit">Marcar entregado</button>
              </form>
            </div>
          {% endfor %}
        </div>
      {% endfor %}
    </div>
  {% else %}
    <div class="empty">✅ No hay ítems pendientes. Todo está entregado.</div>
  {% endif %}
</div>
</body>
</html>
"""



# ============================================================
# Boot
# ============================================================
init_db()

if __name__ == "__main__":
    import os
    # Railway asigna un puerto dinámico, esto lo captura:
    port = int(os.environ.get("PORT", 5000))
    app.run(host='0.0.0.0', port=port)
